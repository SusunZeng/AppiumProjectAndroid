#coding=utf-8
'''
文件名：read_excel.py
作用:引用openpyxl模块来读取excel，其中兼容2007版本以上的excel，含后缀.xlsx文件.与opera_excel.py不同的地方是，opera_excel只兼容.xls文件
作者：曾志坤，时间：20180419
'''
import sys,time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.styles import colors
from openpyxl.styles import Fill,fills
from openpyxl.formatting.rule import ColorScaleRule

class Read_Excel:
    def __init__(self,file_name,sheet_name,):
        '''
        构造函数
        :param file_path:
        '''
        self.filename = file_name
        self.excel = self.get_excel()
        self.table = self.get_table(sheet_name)

    def get_excel(self):
        '''
        获取excel值
        :return:
        '''
        excel = load_workbook(self.filename)
        return excel
    def get_table(self,sheet_name):
        '''
        获取工作表Sheet表的值
        :param sheet_name:
        :return:
        '''
        # table = self.excel.get_sheet_by_name(sheet_name)  #get_sheet_by_name()在调试时，提示该方法已过时
        # table = self.excel.worksheets[0]  #这种方法等同于下方的方法，但为了方便传入参数，使用以下方法
        table = self.excel[sheet_name]
        return table
    def get_rows(self):
        '''
        获取最大行数
        :return:
        '''
        lines = self.table.max_row
        return lines
    def get_cols(self):
        '''
        获取最大列数
        :return:
        '''
        cols = self.table.max_column
        return cols
    def get_cell(self, row, column):
        '''
        获取单元格内容
        :param rows:
        :param columns:
        :return:
        '''
        # 获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
        data=  self.table.cell(row = row, column = column).value
        return data

    def get_excel_value(self, casename):
        rows = self.get_rows()
        # print('rows方法get_excel_value:', rows)
        cols = self.get_cols()
        # print('cols方法get_excel_value:', cols)
        # print('cell_value方法get_excel_value:',self.get_cell(4, 1))
        list = []
        for i in range(1, rows+1):
            # print(self.get_cell(i,0).value)
            if self.get_cell(i, 1) == casename:
                for j in range(1, cols+1):
                    list.append(self.get_cell(i, j))
                return list
            else:
                continue


if __name__ == '__main__':
    read_excel = Read_Excel('E:\\AppiumProjectAndroid\\config\\AndroidAutomationTestCase-1.xlsx','Sheet1')
    print('excel:',read_excel.get_excel())
    print('table:',read_excel.get_table('Sheet1'))
    print('rows:',read_excel.get_rows())
    print('cols:',read_excel.get_cols())
    print('data1:', read_excel.get_cell(4, 10))
    print('data2:', read_excel.get_cell(6, 10))
    print('data2:', read_excel.get_cell(7, 10))
    print('get_excel_value[11]:', read_excel.get_excel_value('执行案例编号')[11])
    print('get_excel_value[3]:', read_excel.get_excel_value('test_3')[9])
    print('get_excel_value[9]:', read_excel.get_excel_value('test_2')[9])


# def get_column_idx(file_path, sheet_name, column_name):
#     wb = openpyxl.load_workbook(file_path)
#     ws = wb.get_sheet_by_name(sheet_name)
#     for i in range(1, ws.max_column+1):
#         if ws.cell(row=1, column=i).value == column_name:
#             break
#     wb.close()
#     return i
#
# def get_case_value(file_path, sheet_name, case_name, column_name):
#     idx = get_column_idx(file_path, sheet_name, column_name)
#     wb = openpyxl.load_workbook(file_path)
#     ws = wb.get_sheet_by_name(sheet_name)
#     for i in range(1, ws.max_row+1):
#         if ws.cell(row=i, column=1).value == case_name:
#             break
#     wb.close()
#     return ws.cell(row=i, column=idx).value
#
# def get_case_list(file_path, sheet_name, case_name):
#     wb = openpyxl.load_workbook(file_path)
#     ws = wb.get_sheet_by_name(sheet_name)
#     for i in range(1, ws.max_row+1):
#         if ws.cell(row=i, column=1).value == case_name:
#             break
#     lists = []
#     for j in range(2, ws.max_column+1):
#         lists.append(ws.cell(row=i, column=j).value)
#     wb.close()
#     return lists


