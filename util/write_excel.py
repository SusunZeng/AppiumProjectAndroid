#coding=utf-8
'''
文件名：write_excel.py
作用：写入数据/设置单元格格式
作者：曾志坤，时间：2018-04-19
参考：https://blog.csdn.net/hqzxsc2006/article/details/51784641
'''

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference, BarChart3D
from openpyxl.styles import Color, Font, Alignment
from openpyxl.styles.colors import BLUE, RED, GREEN, YELLOW


class Write_Excel(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write(self, row, column, value):
        '''
        eg: coord:A1
        self.ws.cell(coord).value = value
        在指定的单元格写入值
        :param row:
        :param column:
        :param value:
        :return:
        '''
        self.ws.cell(row, column, value)
        self.wb.save(self.filename)
        # self.wb.close()

    def merge(self, rangstring):
        '''
        合并单元格
        eg: rangstring:A1:E1
        :param rangstring:
        :return:
        '''
        self.ws.merge_cells(rangstring)
        self.wb.save(self.filename)

    def cellstyle(self, row, column, font, align):
        '''
        单元格加入样式，如字体，颜色等属性
        单元格B2设置宋体，14号，红色，自动换行，水平居中，垂直居中
        :param row:
        :param column:
        :param font:
        :param align:
        :return:
        '''
        cell = self.ws.cell(row, column)
        cell.font = font
        cell.alignment = align

    def makechart(self, title, pos, width, height, col1, row1, col2, row2, col3, row3, row4):
        '''
        创建3d柱状图
        :param title:图表名
        pos:图表位置
        width:图表宽度
        height:图表高度
        官方文档：https://openpyxl.readthedocs.io/en/latest/usage.html
        '''
        data = Reference(self.ws, min_col=col1, min_row=row1, max_col=col2, max_row=row2)
        cat = Reference(self.ws, min_col=col3, min_row=row3, max_row=row4)
        chart = BarChart3D()
        chart.title = title
        chart.width = width
        chart.height = height
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(cat)
        self.ws.add_chart(chart, pos)
        self.wb.save(self.filename)

if __name__ == '__main__':
    wr = Write_Excel('E:\\AppiumProjectAndroid\\config\\test.xlsx')
    # wr.write('A2', 'hello')
    wr.write(3, 4, 'hello')
    wr.write(3, 5, 'boy')
    wr.write(4, 6, 'amazing')
    wr.write(5, 7,'Unbelievable')

    font = Font(name=u'宋体', size=14, color=RED, bold=True)
    align = Alignment(horizontal='center', vertical='center')
    wr.cellstyle(3, 4, font, align)